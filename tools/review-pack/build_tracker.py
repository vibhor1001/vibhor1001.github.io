#!/usr/bin/env python3
"""Build PropertyFlow-Feedback-Tracker.xlsx (v2) — seeded with the audit's open
decisions (OP-01…OP-28) and an Implemented-changes sheet for verification."""
import json
import os
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

SCRATCH = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SCRATCH, "PropertyFlow-Feedback-Tracker.xlsx")

with open(os.path.join(SCRATCH, "pdf", "manifest.json")) as f:
    manifest = json.load(f)
with open(os.path.join(SCRATCH, "pdf", "sheets.json")) as f:
    sheets = json.load(f)
with open(os.path.join(SCRATCH, "audit", "open-items.json")) as f:
    open_items = json.load(f)
entries = [json.loads(l) for l in open(os.path.join(SCRATCH, "audit", "replace-log.jsonl"))]
pages = manifest["pages"]
offsets = sheets["offsets"]

CHARCOAL = "141A23"
ORANGE = "E65A38"
GREY = "8A92A0"
thin = Side(style="thin", color="D9D4CC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=CHARCOAL)
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


# ── How to use ───────────────────────────────────────────────────────
ws = wb.active
ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 130
rows = [
    ("t", "PropertyFlow website feedback tracker — round 2"),
    ("p", "Site under review: https://vibhor1001.github.io/ (58 pages). Read alongside "
          "'PropertyFlow-Website-Review-Pack-v2.pdf' (post-fix render) — the dark banner on each PDF page gives the "
          "part number (P01–P58); the grey footer numbers the sheets within a part."),
    ("s", ""),
    ("h", "What has already happened"),
    ("p", "A full audit raised ~250 findings. Around 120 objective errors are already fixed on the branch — the exact "
          "before/after of every one is in docs/staging-review/changes-log.md and summarised on the 'Implemented changes' "
          "sheet here. Tick them off as you verify each in the v2 PDF."),
    ("p", "28 items need a human decision — they are pre-loaded on the 'Feedback log' sheet as OP-01…OP-28. "
          "Add your own findings underneath as FB-001, FB-002, …"),
    ("s", ""),
    ("h", "Logging your own feedback"),
    ("p", "One row per issue: part + sheet from the PDF, the exact current text (copy from the PDF — text is selectable), "
          "your proposed text, an issue type and a priority. Leave Status = 'New'."),
    ("p", "Record wording rulings (user vs landlord vs host, short-term let vs short-let vs STL, plan names …) on the "
          "'Terminology' sheet — the developer applies those site-wide."),
    ("s", ""),
    ("h", "The loop (nothing goes live)"),
    ("p", "Send this file + your annotated PDF back by email/shared drive (not GitHub). The developer works the rows, "
          "marks them Implemented, and exports a fresh pack for you to verify. Everything stays on the review branch — "
          "publishing to the live site only happens after every row is Verified/Rejected and you sign off."),
]
r = 1
for kind, text in rows:
    cell = ws.cell(row=r, column=2, value=text)
    if kind == "t":
        cell.font = Font(size=18, bold=True, color=CHARCOAL)
        ws.row_dimensions[r].height = 30
    elif kind == "h":
        cell.font = Font(size=13, bold=True, color=ORANGE)
        ws.row_dimensions[r].height = 22
    else:
        cell.font = Font(size=11, color="333333")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if len(text) > 110:
            ws.row_dimensions[r].height = 42
    r += 1

# ── Feedback log ─────────────────────────────────────────────────────
ws = wb.create_sheet("Feedback log")
headers = [
    ("Ref", 8), ("Page (part)", 11), ("Sheet #", 7), ("Where / pages affected", 30),
    ("Issue — current state", 52), ("Issue type", 18), ("Proposed change / action needed", 46),
    ("Priority", 12), ("Status", 13), ("Owner", 12), ("Developer notes", 28),
]
for i, (h, w) in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, len(headers))
ws.freeze_panes = "A2"

TYPE_BY_PRIORITY_HINT = {
    "OP-02": "Legal / compliance", "OP-18": "Legal / compliance", "OP-19": "Legal / compliance",
    "OP-04": "Terminology", "OP-05": "Content accuracy", "OP-06": "Content accuracy",
    "OP-07": "Content accuracy", "OP-08": "Content accuracy", "OP-10": "Content accuracy",
    "OP-03": "Technical (dev)", "OP-17": "Technical (dev)", "OP-25": "Technical (dev)",
    "OP-26": "Technical (dev)", "OP-28": "Technical (dev)", "OP-23": "Technical (dev)",
}
row_i = 2
for it in open_items:
    vals = [it["ref"], "—", "", it["pages"], it["finding"],
            TYPE_BY_PRIORITY_HINT.get(it["ref"], "Wording & tone" if it["priority"] == "Nice to have" else "Content accuracy"),
            it["action"], it["priority"], "New", "", ""]
    for j, v in enumerate(vals, 1):
        ws.cell(row=row_i, column=j, value=v)
    row_i += 1

N_ROWS = 320
for rr in range(2, N_ROWS + 2):
    for cc in range(1, len(headers) + 1):
        cell = ws.cell(row=rr, column=cc)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border

dv_part = DataValidation(type="list", formula1="PartList", allow_blank=True)
dv_type = DataValidation(type="list", allow_blank=True, formula1=
    '"Content accuracy,Wording & tone,Terminology,Spelling & grammar,CTA / button copy,Legal / compliance,Technical (dev),Other"')
dv_prio = DataValidation(type="list", allow_blank=True, formula1='"Must fix,Should fix,Nice to have"')
dv_stat = DataValidation(type="list", allow_blank=True, formula1='"New,Agreed,In progress,Implemented,Verified,Rejected"')
for dv, col in [(dv_part, "B"), (dv_type, "F"), (dv_prio, "H"), (dv_stat, "I")]:
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{N_ROWS + 1}")
ws.auto_filter.ref = f"A1:K{N_ROWS + 1}"

# ── Implemented changes ──────────────────────────────────────────────
ws = wb.create_sheet("Implemented changes")
headers = [("Change id", 24), ("Category", 12), ("Hits", 7), ("Files", 34),
           ("Was", 48), ("Now", 48), ("Verified in v2 PDF?", 18)]
for i, (h, w) in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, len(headers))
ws.freeze_panes = "A2"
by_id = OrderedDict()
for e in entries:
    k = e["id"]
    if k not in by_id:
        by_id[k] = {"category": e["category"], "find": e["find"], "replace": e["replace"], "files": [], "count": 0}
    by_id[k]["files"].append(e["file"])
    by_id[k]["count"] += e["count"]
for k, v in by_id.items():
    files = ", ".join(v["files"][:4]) + (f" … +{len(v['files'])-4}" if len(v["files"]) > 4 else "")
    ws.append([k, v["category"], v["count"], files, v["find"], v["replace"], ""])
for rr in range(2, ws.max_row + 1):
    for cc in range(1, len(headers) + 1):
        cell = ws.cell(row=rr, column=cc)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
dv_ver = DataValidation(type="list", allow_blank=True, formula1='"Yes,No,Reopen"')
ws.add_data_validation(dv_ver)
dv_ver.add(f"G2:G{ws.max_row}")
ws.auto_filter.ref = f"A1:G{ws.max_row}"

# ── Pages ────────────────────────────────────────────────────────────
ws = wb.create_sheet("Pages")
headers = [("Part", 8), ("Section", 18), ("Page", 30), ("Path", 40),
           ("Staging URL", 52), ("Starts at PDF page", 17), ("Sheets", 8), ("<title> tag", 60)]
for i, (h, w) in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, len(headers))
ws.freeze_panes = "A2"
ws.append(["Global", "—", "Applies site-wide", "—", "https://vibhor1001.github.io/", "", "", ""])
for p in pages:
    off, n = offsets.get(p["part"], (None, None))
    ws.append([p["part"], p["group"], p["label"], p["path"],
               "https://vibhor1001.github.io" + p["path"],
               (off + 1) if off is not None else "", n or "", p.get("title", "")])
for rr in range(2, ws.max_row + 1):
    for cc in range(1, len(headers) + 1):
        ws.cell(row=rr, column=cc).border = border
wb.defined_names.add(DefinedName("PartList", attr_text=f"Pages!$A$2:$A${ws.max_row}"))

# ── Terminology ──────────────────────────────────────────────────────
ws = wb.create_sheet("Terminology")
headers = [("Preferred term (use this)", 34), ("Variants seen on the site", 44),
           ("Where noticed", 26), ("Ruling / notes", 44)]
for i, (h, w) in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, len(headers))
ws.freeze_panes = "A2"
for row in [
    ["", "short-term let / short-let / STL / STR", "Hero copy, titles, blog", "Coordinator to decide (STR already normalised to STL in several titles)"],
    ["", "landlord / host / user / operator — nav 'For Users' vs pages 'For Landlords' vs nav 'Solutions'", "Site-wide (see OP-04)", "Coordinator to decide"],
    ["", "Free / Premium plan names (was 'Get Started' / 'Full Management')", "Now consistent — confirm you like the names", "Applied per pricing screenshot"],
    ["", "changeovers vs turnovers", "Partner pages vs platform feature 'Turnovers & Team'", "Coordinator to decide"],
    ["", "Smart Pricing / AI pricing / AI Smart Pricing", "Platform + nav ('Cognito Pricing' already retired)", "Coordinator to decide"],
    ["", "council vs local authority", "Blog + resources nav", "Coordinator to decide"],
]:
    ws.append(row)
for rr in range(2, 40):
    for cc in range(1, len(headers) + 1):
        cell = ws.cell(row=rr, column=cc)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border

wb.save(OUT)
print("wrote", OUT, f"({os.path.getsize(OUT)/1024:.0f} KB) — {len(open_items)} open items, {len(by_id)} implemented groups")
